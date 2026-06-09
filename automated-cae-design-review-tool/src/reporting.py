from pathlib import Path
import pandas as pd


def write_markdown_report(summary: pd.DataFrame, block_summary: pd.DataFrame, recommendation: str, output_path: str | Path) -> None:
    output_path = Path(output_path)
    lines = []
    lines.append('# Automated CAE Design Review Report')
    lines.append('')
    lines.append('## Executive Summary')
    lines.append('')
    lines.append(recommendation)
    lines.append('')
    lines.append('## Case Judgement')
    lines.append('')
    cols = ['case_id', 'design_status', 'max_temperature_c', 'max_stress_vm_mpa', 'max_displacement_mm', 'minimum_safety_factor', 'temperature_exceeded_area_percent']
    lines.append(summary[cols].round(3).to_markdown(index=False))
    lines.append('')
    lines.append('## Decision Comments')
    lines.append('')
    for _, row in summary.iterrows():
        lines.append(f"- **{row['case_id']}**: {row['design_status']} - {row['decision_comment']}")
    lines.append('')
    lines.append('## Block-Averaged Heat Flux Check')
    lines.append('')
    lines.append(block_summary.round(4).to_markdown(index=False))
    lines.append('')
    lines.append('## Notes')
    lines.append('')
    lines.append('- Input data in this repository is synthetic and is not derived from confidential CAE data.')
    lines.append('- Replace the criteria in `config/design_criteria.json` with project-specific limits before using this workflow in actual design work.')
    output_path.write_text('\n'.join(lines), encoding='utf-8')


def write_excel_report(summary: pd.DataFrame, block_summary: pd.DataFrame, risk_elements: pd.DataFrame, output_path: str | Path) -> None:
    output_path = Path(output_path)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        summary.to_excel(writer, sheet_name='Case Summary', index=False)
        block_summary.to_excel(writer, sheet_name='Heat Flux Block Check', index=False)
        risk_elements.to_excel(writer, sheet_name='Risk Elements', index=False)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            for cell in ws[1]:
                cell.style = 'Headline 3'
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_length + 2, 12), 42)
        # Add simple conditional formatting for status cells
        from openpyxl.styles import PatternFill
        status_fill = {
            'OK': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'Warning': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'NG': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        }
        ws = wb['Case Summary']
        headers = [c.value for c in ws[1]]
        if 'design_status' in headers:
            status_col = headers.index('design_status') + 1
            for row in range(2, ws.max_row + 1):
                value = ws.cell(row, status_col).value
                if value in status_fill:
                    ws.cell(row, status_col).fill = status_fill[value]
